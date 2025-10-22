#!/usr/bin/env python3
"""
분류 수정 스크립트

사용자 요청사항에 따라 메인 시트 일괄 수정
"""
import openpyxl
import pandas as pd
from pathlib import Path


def apply_corrections():
    """메인 시트 수정 적용"""

    file_path = "output/칠칠기업_법인카드_완성본.xlsx"

    print("="*60)
    print("분류 수정 적용 시작")
    print("="*60)

    # 백업 생성
    backup_path = "output/칠칠기업_법인카드_백업.xlsx"
    import shutil
    shutil.copy(file_path, backup_path)
    print(f"\n✅ 백업 생성: {backup_path}\n")

    # Excel 로드
    wb = openpyxl.load_workbook(file_path)

    total_modified = 0

    # === 4985 카드 수정 ===
    print("[4985 카드 수정]")
    sheet_4985 = wb['4985']
    count_4985 = 0

    for row_idx, row in enumerate(sheet_4985.iter_rows(min_row=2), start=2):
        category_cell = row[3]  # 사용용도
        old_category = category_cell.value

        if old_category in ['복리후생', '복리후생비', '중식대', '기타']:
            category_cell.value = '거래처 교제비'
            count_4985 += 1

            merchant = row[1].value
            amount = row[2].value
            print(f"  Row {row_idx}: {merchant} ({amount:,}원)")
            print(f"    {old_category} → 거래처 교제비")

    print(f"  → {count_4985}건 수정\n")
    total_modified += count_4985

    # === 6974 카드 수정 ===
    print("[6974 카드 수정]")
    sheet_6974 = wb['6974']
    count_6974 = 0

    for row_idx, row in enumerate(sheet_6974.iter_rows(min_row=2), start=2):
        merchant_cell = row[1]  # 가맹점명
        category_cell = row[3]  # 사용용도

        merchant = str(merchant_cell.value) if merchant_cell.value else ""
        old_category = category_cell.value
        new_category = None

        # 1. 정확 일치 규칙 (SKT)
        if merchant == "SKT-자동납부-647179":
            new_category = "휴대폰(010-7504-4043)사용료"
        elif merchant == "SKT-자동납부-044043":
            new_category = "휴대폰(010-3664-7179)사용료"

        # 2. 키워드 규칙
        elif "쿠팡" in merchant:
            new_category = "기숙사 물품 구입비"
        elif "홈플러스" in merchant:
            new_category = "기숙사 물품 구입비"
        elif "트레이더스" in merchant:
            new_category = "기숙사 물품 구입비"

        if new_category and new_category != old_category:
            category_cell.value = new_category
            count_6974 += 1

            amount = row[2].value
            print(f"  Row {row_idx}: {merchant} ({amount:,}원)")
            print(f"    {old_category} → {new_category}")

    print(f"  → {count_6974}건 수정\n")
    total_modified += count_6974

    # 저장
    wb.save(file_path)

    print("="*60)
    print(f"총 {total_modified}건 수정 완료")
    print("="*60)

    # 검증
    print("\n[수정 후 검증]\n")

    # 4985 카드 검증
    df_4985 = pd.read_excel(file_path, sheet_name='4985')
    old_categories = df_4985[df_4985['사용용도'].isin(['복리후생', '중식대', '기타'])]
    print(f"4985: 복리후생/중식대/기타 남은 건수 = {len(old_categories)}건 (0이어야 함)")

    new_categories = df_4985[df_4985['사용용도'] == '거래처 교제비']
    print(f"4985: 거래처 교제비 건수 = {len(new_categories)}건")

    # 6974 카드 검증
    df_6974 = pd.read_excel(file_path, sheet_name='6974')
    coupang = df_6974[df_6974['가맹점명'].str.contains('쿠팡', na=False)]
    dormitory = coupang[coupang['사용용도'] == '기숙사 물품 구입비']
    print(f"\n6974: 쿠팡 → 기숙사 물품 구입비 = {len(dormitory)}/{len(coupang)}건")

    skt_647179 = df_6974[df_6974['가맹점명'] == 'SKT-자동납부-647179']
    print(f"6974: SKT-647179 → 휴대폰(010-7504-4043)사용료 = {len(skt_647179)}건")

    skt_044043 = df_6974[df_6974['가맹점명'] == 'SKT-자동납부-044043']
    print(f"6974: SKT-044043 → 휴대폰(010-3664-7179)사용료 = {len(skt_044043)}건")

    print("\n✅ 수정 완료!")


if __name__ == "__main__":
    apply_corrections()
